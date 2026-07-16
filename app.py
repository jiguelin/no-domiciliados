import streamlit as st
import anthropic
import base64
import json
import re
import io
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
import os

# ── Configuración de página ──────────────────────────────────────────────────
st.set_page_config(
    page_title="Facturas del Exterior",
    page_icon="🧾",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── CSS personalizado ─────────────────────────────────────────────────────────
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(90deg, #1e3a5f 0%, #2d6a9f 100%);
        padding: 1.5rem 2rem;
        border-radius: 10px;
        margin-bottom: 1.5rem;
        color: white;
    }
    .main-header h1 { color: white; margin: 0; font-size: 1.8rem; }
    .main-header p { color: #b8d4f0; margin: 0.3rem 0 0; font-size: 0.95rem; }
    .step-card {
        background: #f8fafc;
        border: 1px solid #e2e8f0;
        border-left: 4px solid #2d6a9f;
        border-radius: 8px;
        padding: 1rem 1.2rem;
        margin-bottom: 1rem;
    }
    .step-card h3 { color: #1e3a5f; margin: 0 0 0.5rem; font-size: 1rem; }
    .success-box {
        background: #f0fdf4;
        border: 1px solid #86efac;
        border-radius: 8px;
        padding: 0.8rem 1rem;
        color: #166534;
    }
    .warning-box {
        background: #fffbeb;
        border: 1px solid #fcd34d;
        border-radius: 8px;
        padding: 0.8rem 1rem;
        color: #92400e;
    }
    .info-box {
        background: #eff6ff;
        border: 1px solid #93c5fd;
        border-radius: 8px;
        padding: 0.8rem 1rem;
        color: #1e40af;
    }
    .metric-card {
        background: white;
        border: 1px solid #e2e8f0;
        border-radius: 8px;
        padding: 1rem;
        text-align: center;
    }
    .metric-card .value { font-size: 1.6rem; font-weight: 700; color: #1e3a5f; }
    .metric-card .label { font-size: 0.8rem; color: #64748b; margin-top: 0.2rem; }
    .empresa-tag {
        display: inline-block;
        background: #dbeafe;
        color: #1e40af;
        border-radius: 12px;
        padding: 2px 10px;
        font-size: 0.8rem;
        font-weight: 600;
    }
    div[data-testid="stDataFrame"] { border-radius: 8px; overflow: hidden; }
    .stButton > button {
        border-radius: 8px;
        font-weight: 600;
    }
</style>
""", unsafe_allow_html=True)

# ── API Key ───────────────────────────────────────────────────────────────────
API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
if not API_KEY:
    st.error("⚠️ No se encontró la API key de Anthropic. Configúrala en Streamlit Secrets como `ANTHROPIC_API_KEY`.")
    st.stop()

client = anthropic.Anthropic(api_key=API_KEY)

# ── Session state ─────────────────────────────────────────────────────────────
if "tipo_cambio" not in st.session_state:
    st.session_state.tipo_cambio = {}
if "filas" not in st.session_state:
    st.session_state.filas = []
if "empresa_activa" not in st.session_state:
    st.session_state.empresa_activa = "EMPRESA1"
if "mes_activo" not in st.session_state:
    st.session_state.mes_activo = ""
if "nps_cruzados" not in st.session_state:
    st.session_state.nps_cruzados = {}

# ── Helpers ───────────────────────────────────────────────────────────────────
def pdf_to_base64(file_bytes: bytes) -> str:
    return base64.standard_b64encode(file_bytes).decode("utf-8")

def get_tc(fecha_str: str, tc_dict: dict) -> float:
    """Devuelve el tipo de cambio venta para una fecha dada."""
    try:
        d = datetime.strptime(fecha_str, "%Y-%m-%d")
        day = d.day
        return tc_dict.get(day, 0.0)
    except:
        return 0.0

def calcular_fila(usd: float, tc: float) -> dict:
    sin_igv = round(usd * tc, 2)
    igv_soles = round(sin_igv * 0.18, 2)
    igv_usd = round(usd * 0.18, 2)
    total_soles = round(sin_igv + igv_soles, 2)
    renta = round(sin_igv * 0.30, 2)
    renta_usd = round(usd * 0.30, 2)
    return {
        "sin_igv": sin_igv,
        "igv_soles": igv_soles,
        "igv_usd": igv_usd,
        "total_soles": total_soles,
        "renta": renta,
        "renta_usd": renta_usd,
    }

# ── Extracción con IA ─────────────────────────────────────────────────────────
def extraer_tipo_cambio(pdf_bytes: bytes) -> dict:
    """Lee el PDF de tipo de cambio SUNAT y devuelve {dia: tc_venta}."""
    b64 = pdf_to_base64(pdf_bytes)
    prompt = """Extrae el tipo de cambio VENTA de cada día del mes de este PDF de SUNAT.
Responde SOLO con JSON válido sin markdown, formato exacto:
{"1": 3.495, "2": 3.467, ...}
Solo incluye los días que tienen valor. Usa punto decimal."""
    
    resp = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=1000,
        messages=[{
            "role": "user",
            "content": [
                {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": b64}},
                {"type": "text", "text": prompt}
            ]
        }]
    )
    text = resp.content[0].text.strip()
    text = re.sub(r"```json|```", "", text).strip()
    raw = json.loads(text)
    return {int(k): float(v) for k, v in raw.items()}

def extraer_invoice(pdf_bytes: bytes, tc_dict: dict, empresa: str) -> dict | None:
    """Extrae datos de un PDF de invoice usando Claude."""
    b64 = pdf_to_base64(pdf_bytes)
    prompt = f"""Extrae los datos de esta factura del exterior. Responde SOLO con JSON válido sin markdown:
{{
  "empresa": "{empresa}",
  "fecha": "YYYY-MM-DD",
  "proveedor": "nombre completo del emisor",
  "pais": "país del emisor en español",
  "domicilio": "dirección completa del emisor",
  "vat": "número VAT/RUC/Tax ID del emisor",
  "invoice": "número de factura exacto",
  "monto_original": número_decimal_del_monto_tal_como_aparece_en_la_factura,
  "moneda_original": "USD o PEN"
}}
IMPORTANTE:
- El campo monto_original es el monto EXACTO como aparece en la factura (puede ser USD o PEN/soles).
- El campo moneda_original indica la moneda: "USD" si está en dólares, "PEN" si está en soles peruanos.
- NO conviertas monedas, solo extrae el monto original y la moneda."""

    resp = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=1000,
        messages=[{
            "role": "user",
            "content": [
                {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": b64}},
                {"type": "text", "text": prompt}
            ]
        }]
    )
    text = resp.content[0].text.strip()
    text = re.sub(r"```json|```", "", text).strip()
    try:
        data = json.loads(text)
        monto_original = float(data.get("monto_original", 0))
        moneda = data.get("moneda_original", "USD").upper()
        fecha = data.get("fecha", "")
        tc = get_tc(fecha, tc_dict)

        # Conversión PEN→USD hecha en Python (más confiable que pedirle a la IA)
        if moneda == "PEN":
            usd = round(monto_original / tc, 2) if tc > 0 else 0.0
        else:
            usd = monto_original

        calcs = calcular_fila(usd, tc)
        return {
            "empresa": empresa,
            "fecha": fecha,
            "proveedor": data.get("proveedor", ""),
            "pais": data.get("pais", ""),
            "domicilio": data.get("domicilio", ""),
            "vat": data.get("vat", ""),
            "invoice": data.get("invoice", ""),
            "nac": "16630-",
            "usd": usd,
            "tc": tc,
            **calcs,
            "estado": "OK",
        }
    except Exception as e:
        st.warning(f"No se pudo procesar un invoice: {e}")
        return None

def extraer_detalle_declaraciones(file_bytes: bytes, filename: str) -> list:
    """
    Extrae las filas de pago del reporte SUNAT 'Detalle de Declaraciones y Pagos'.
    Acepta PDF o Excel (.xls/.xlsx) y devuelve una lista de:
    {"periodo": "202605", "n_orden": "228769105", "importe": 276}
    Solo se consideran filas tipo 'BOLETA DE PAGO - NPS' (pagos reales),
    excluyendo PLANILLA ELECTRONICA y PDT OTRAS RETENCIONES con importe 0.
    """
    ext = filename.lower().split(".")[-1]

    if ext in ("xls", "xlsx"):
        df = None
        for engine in (["xlrd"] if ext == "xls" else ["openpyxl"]):
            try:
                df = pd.read_excel(io.BytesIO(file_bytes), engine=engine, header=None)
                break
            except Exception:
                continue
        if df is None:
            try:
                df = pd.read_excel(io.BytesIO(file_bytes), header=None)
            except Exception as e:
                raise ValueError(f"No se pudo leer el archivo Excel: {e}")

        # Localizar la fila de cabecera real (la que contiene "N°ORDEN")
        header_row_idx = None
        for i in range(len(df)):
            fila_vals = [str(v).upper() for v in df.iloc[i].tolist()]
            if any(("N°ORDEN" in v) or ("NORDEN" in v.replace("°", "")) for v in fila_vals):
                header_row_idx = i
                break

        pagos = []
        if header_row_idx is not None:
            cols = df.iloc[header_row_idx].astype(str).str.upper().str.strip().tolist()

            def find_col(*keywords):
                for idx, c in enumerate(cols):
                    if all(k in c for k in keywords):
                        return idx
                return None

            col_periodo = find_col("PERIODO")
            col_orden = find_col("ORDEN")
            col_desc = find_col("DESCRIPCION")
            col_importe = find_col("IMPORTE")

            for i in range(header_row_idx + 1, len(df)):
                row = df.iloc[i]
                desc_val = str(row[col_desc]) if col_desc is not None else ""
                if "BOLETA DE PAGO" not in desc_val.upper() and "NPS" not in desc_val.upper():
                    continue
                try:
                    importe_raw = row[col_importe] if col_importe is not None else None
                    importe = int(round(float(importe_raw)))
                except (TypeError, ValueError):
                    continue
                if importe <= 0:
                    continue
                n_orden = str(row[col_orden]).strip() if col_orden is not None else ""
                n_orden = re.sub(r"\.0$", "", n_orden)
                periodo = str(row[col_periodo]).strip() if col_periodo is not None else ""
                periodo = re.sub(r"\.0$", "", periodo)
                pagos.append({"periodo": periodo, "n_orden": n_orden, "importe": importe})

        if pagos:
            return pagos
        # Si el parseo estructurado no encontró nada, cae a extracción con IA como respaldo
        tabla_texto = df.to_csv(index=False)
        content_block = {"type": "text", "text": f"Contenido del reporte SUNAT (filas crudas):\n{tabla_texto}"}
    else:
        b64 = pdf_to_base64(file_bytes)
        content_block = {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": b64}}

    prompt = """Este es el reporte SUNAT "Detalle de Declaraciones y Pagos". Extrae SOLO las filas cuya
DESCRIPCION sea "BOLETA DE PAGO - NPS" (son los pagos reales de IGV/renta no domiciliado).
Ignora filas como "PLANILLA ELECTRONICA" o "PDT OTRAS RETENCIONES" o cualquier fila con IMPORTE PAGADO igual a 0.
Para cada fila relevante extrae: PERIODO, N°ORDEN, e IMPORTE PAGADO (siempre un número entero en soles, sin decimales).

Responde SOLO con JSON válido sin markdown, en este formato exacto:
[
  {"periodo": "202605", "n_orden": "228769105", "importe": 276},
  {"periodo": "202605", "n_orden": "228768591", "importe": 2}
]"""

    resp = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=2000,
        messages=[{"role": "user", "content": [content_block, {"type": "text", "text": prompt}]}]
    )
    text = resp.content[0].text.strip()
    text = re.sub(r"```json|```", "", text).strip()
    try:
        data = json.loads(text)
        for d in data:
            d["importe"] = int(round(float(d["importe"])))
        return data
    except Exception:
        return []


def extraer_formulario_1662(pdf_bytes: bytes) -> list:
    b64 = pdf_to_base64(pdf_bytes)
    prompt = """Este es un PDF del 'Resumen de Transacciones' o constancias del Formulario 1662 de SUNAT
para pago de IGV servicios no domiciliados. Extrae cada pago individual.
Para cada uno extrae el Numero de Orden, el numero de comprobante (Serie y numero),
y el importe pagado en soles.
El numero de comprobante aparece en formatos como:
- '62A8E305 - 0009' -> comprobante: '62A8E305-0009'
- '702 - 106054117' -> comprobante: '702-106054117'
- '379 - 100001292' -> comprobante: '379-100001292'
- 'INV - 5626352153' -> comprobante: 'INV-5626352153'
Responde SOLO con JSON valido sin markdown:
[
  {"n_orden": "1196817219", "comprobante": "62A8E305-0009", "importe": 14},
  {"n_orden": "1196817220", "comprobante": "702-106054117", "importe": 274}
]
El campo comprobante une serie y numero con guion, sin espacios.
El importe siempre es numero entero en soles."""
    resp = client.messages.create(
        model="claude-sonnet-4-6", max_tokens=2000,
        messages=[{"role": "user", "content": [
            {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": b64}},
            {"type": "text", "text": prompt}
        ]}]
    )
    text = resp.content[0].text.strip()
    text = re.sub(r"```json|```", "", text).strip()
    try:
        data = json.loads(text)
        for d in data:
            d["importe"] = int(round(float(d["importe"])))
        return data
    except Exception:
        return []

def cruzar_formulario_1662_con_invoices(pagos_1662: list, filas: list) -> dict:
    resultado = {}
    filas_pendientes = list(range(len(filas)))
    for pago in pagos_1662:
        # Normalizar comprobante: quitar guiones, espacios, prefijos como INV
        comp_raw = pago.get("comprobante", "").replace(" ", "").replace("-", "").upper()
        comp_raw = re.sub(r"^INV", "", comp_raw)  # quitar prefijo INV de Google
        for idx in filas_pendientes:
            invoice = str(filas[idx].get("invoice", "")).replace(" ", "").replace("-", "").upper()
            invoice = re.sub(r"^INV", "", invoice)
            # Match si uno contiene al otro o son iguales
            if comp_raw and invoice and (comp_raw in invoice or invoice.endswith(comp_raw) or invoice == comp_raw or comp_raw.endswith(invoice)):
                resultado[idx] = pago["n_orden"]
                filas_pendientes.remove(idx)
                break
    return resultado

def calcular_renta_por_proveedor(filas: list) -> list:
    from collections import defaultdict
    grupos = defaultdict(lambda: {"proveedor": "", "pais": "", "vat": "", "sin_igv": 0.0, "renta": 0.0, "invoices": 0})
    for f in filas:
        key = f.get("vat", f.get("proveedor", ""))
        grupos[key]["proveedor"] = f.get("proveedor", "")
        grupos[key]["pais"] = f.get("pais", "")
        grupos[key]["vat"] = f.get("vat", "")
        grupos[key]["sin_igv"] += float(f.get("sin_igv", 0))
        grupos[key]["renta"] += float(f.get("renta", 0))
        grupos[key]["invoices"] += 1
    result = []
    for g in grupos.values():
        result.append({
            "Proveedor": g["proveedor"], "País": g["pais"], "RUC/VAT": g["vat"],
            "N° invoices": g["invoices"],
            "Total S/. sin IGV": round(g["sin_igv"], 2),
            "Renta 30% S/.": round(g["renta"], 2),
        })
    return sorted(result, key=lambda x: x["Total S/. sin IGV"], reverse=True)


def cruzar_pagos_con_invoices(pagos: list, filas: list) -> dict:
    """
    Cruza los pagos SUNAT (montos redondos en soles) con los invoices por monto de IGV redondeado.
    Si dos invoices tienen el mismo IGV redondeado, el orden es intercambiable (mismo resultado).
    Retorna {indice_fila: numero_orden}
    """
    resultado = {}
    filas_pendientes = list(range(len(filas)))

    for pago in pagos:
        monto_pago = int(round(pago["importe"]))

        for idx in filas_pendientes:
            fila = filas[idx]
            igv_fila = int(round(fila["igv_soles"]))
            if igv_fila == monto_pago:
                resultado[idx] = pago["n_orden"]
                filas_pendientes.remove(idx)
                break

    return resultado

# ── Generación de Excel ───────────────────────────────────────────────────────
def generar_excel(filas: list, mes: str) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    empresas = list(dict.fromkeys(f["empresa"] for f in filas))

    # Estilos
    header_fill = PatternFill("solid", fgColor="1e3a5f")
    header_font = Font(bold=True, color="FFFFFF", size=9)
    total_fill = PatternFill("solid", fgColor="dbeafe")
    total_font = Font(bold=True, size=9)
    data_font = Font(size=9)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    headers = [
        "Fecha Invoice", "Proveedor", "PAIS DE RESIDENCIA", "DOMICILIO",
        "RUC/VAT", "Invoice", "Nacionalización", "Monto USD",
        "Tipo de cambio\nSUNAT (venta)", "Monto en S/.\n(sin IGV)",
        "IGV (18%)\nen S/.", "IGV USD", "Total con\nIGV (S/.)",
        "Renta", "Renta USD"
    ]
    col_widths = [12, 28, 16, 35, 14, 22, 18, 10, 11, 13, 11, 10, 13, 11, 10]

    for empresa in empresas:
        ws = wb.create_sheet(title=f"{empresa}_{mes}")
        rows_emp = [f for f in filas if f["empresa"] == empresa]

        # Cabecera
        for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
            ws.column_dimensions[get_column_letter(col_idx)].width = w
        ws.row_dimensions[1].height = 30

        # Datos
        for row_idx, fila in enumerate(rows_emp, 2):
            fecha_val = fila["fecha"]
            try:
                fecha_val = datetime.strptime(fila["fecha"], "%Y-%m-%d")
            except:
                pass

            valores = [
                fecha_val, fila["proveedor"], fila["pais"], fila["domicilio"],
                fila["vat"], fila["invoice"], fila.get("nac", "16630-"),
                fila["usd"], fila["tc"], fila["sin_igv"],
                fila["igv_soles"], fila["igv_usd"], fila["total_soles"],
                fila["renta"], fila["renta_usd"]
            ]

            for col_idx, val in enumerate(valores, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = data_font
                cell.border = border
                if col_idx == 1 and isinstance(val, datetime):
                    cell.number_format = "DD/MM/YYYY"
                    cell.alignment = center
                elif col_idx in [8, 9, 10, 11, 12, 13, 14, 15]:
                    cell.number_format = "#,##0.00"
                    cell.alignment = right
                else:
                    cell.alignment = Alignment(vertical="center", wrap_text=True)

        # Fila de totales
        tot_row = len(rows_emp) + 2
        ws.cell(row=tot_row, column=1, value="Totales").font = total_font
        ws.cell(row=tot_row, column=2, value="—").font = total_font
        for col_idx in range(1, 16):
            ws.cell(row=tot_row, column=col_idx).fill = total_fill
            ws.cell(row=tot_row, column=col_idx).border = border

        sum_cols = {8: "usd", 10: "sin_igv", 11: "igv_soles", 13: "total_soles", 14: "renta"}
        for col_idx, field in sum_cols.items():
            val = sum(f[field] for f in rows_emp)
            cell = ws.cell(row=tot_row, column=col_idx, value=round(val, 2))
            cell.font = total_font
            cell.number_format = "#,##0.00"
            cell.alignment = right

        ws.freeze_panes = "A2"

    # ── Pestaña de Renta por proveedor ───────────────────────────────────────
    ws_renta = wb.create_sheet(title=f"RENTA_{mes}")
    renta_headers = ["Proveedor", "País", "RUC/VAT", "N° invoices", "Total S/. sin IGV", "Renta 30% S/."]
    renta_widths = [30, 16, 16, 12, 20, 18]

    for col_idx, (h, w) in enumerate(zip(renta_headers, renta_widths), 1):
        cell = ws_renta.cell(row=1, column=col_idx, value=h)
        cell.fill = PatternFill("solid", fgColor="0f4c35")
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=Side(style="thin", color="CBD5E1"), right=Side(style="thin", color="CBD5E1"),
                             top=Side(style="thin", color="CBD5E1"), bottom=Side(style="thin", color="CBD5E1"))
        ws_renta.column_dimensions[get_column_letter(col_idx)].width = w
    ws_renta.row_dimensions[1].height = 24

    renta_data = calcular_renta_por_proveedor(filas)
    for row_idx, r in enumerate(renta_data, 2):
        for col_idx, key in enumerate(renta_headers, 1):
            val = r[key]
            cell = ws_renta.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(size=9)
            cell.border = Border(left=Side(style="thin", color="CBD5E1"), right=Side(style="thin", color="CBD5E1"),
                                 top=Side(style="thin", color="CBD5E1"), bottom=Side(style="thin", color="CBD5E1"))
            if col_idx in (5, 6):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx == 4:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Fila total renta
    tot_renta_row = len(renta_data) + 2
    ws_renta.cell(row=tot_renta_row, column=1, value="TOTAL").font = Font(bold=True, size=9)
    for col_idx in range(1, 7):
        ws_renta.cell(row=tot_renta_row, column=col_idx).fill = PatternFill("solid", fgColor="d1fae5")
        ws_renta.cell(row=tot_renta_row, column=col_idx).border = Border(
            left=Side(style="thin", color="CBD5E1"), right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"), bottom=Side(style="thin", color="CBD5E1"))
    for col_idx, key in [(5, "Total S/. sin IGV"), (6, "Renta 30% S/.")]:
        val = sum(r[key] for r in renta_data)
        cell = ws_renta.cell(row=tot_renta_row, column=col_idx, value=round(val, 2))
        cell.font = Font(bold=True, size=9)
        cell.number_format = "#,##0.00"
        cell.alignment = Alignment(horizontal="right", vertical="center")
    ws_renta.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

# ── Reporte de IGV / NPS pendientes ───────────────────────────────────────────
def filas_nps_pendientes(filas: list) -> list:
    """Filtra los invoices con IGV > S/1 (los que requieren generar NPS individual)."""
    pendientes = []
    for f in filas:
        igv_redondeado = int(round(f["igv_soles"]))
        if igv_redondeado > 1:
            pendientes.append({**f, "igv_redondeado": igv_redondeado})
    return pendientes

def generar_excel_nps(filas: list, mes: str) -> bytes:
    """Genera un Excel simple con los datos necesarios para crear los NPS en SUNAT."""
    pendientes = filas_nps_pendientes(filas)

    wb = Workbook()
    ws = wb.active
    ws.title = "NPS_Pendientes"

    header_fill = PatternFill("solid", fgColor="1e3a5f")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    data_font = Font(size=10)
    total_fill = PatternFill("solid", fgColor="dbeafe")
    total_font = Font(bold=True, size=10)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    headers = ["Empresa", "Fecha", "Proveedor", "N° Invoice", "Monto USD", "IGV S/. (a pagar en SUNAT)"]
    widths = [14, 12, 28, 22, 12, 22]

    for col_idx, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[1].height = 24

    row_idx = 2
    for f in pendientes:
        fecha_val = f["fecha"]
        try:
            fecha_val = datetime.strptime(f["fecha"], "%Y-%m-%d")
        except Exception:
            pass
        valores = [f["empresa"], fecha_val, f["proveedor"], f["invoice"], f["usd"], f["igv_redondeado"]]
        for col_idx, val in enumerate(valores, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = border
            if col_idx == 2 and isinstance(val, datetime):
                cell.number_format = "DD/MM/YYYY"
                cell.alignment = center
            elif col_idx in (5, 6):
                cell.number_format = "#,##0.00" if col_idx == 5 else "#,##0"
                cell.alignment = right
            else:
                cell.alignment = Alignment(vertical="center")
        row_idx += 1

    ws.cell(row=row_idx, column=1, value="Total a pagar en SUNAT (IGV)").font = total_font
    for c in range(1, 7):
        ws.cell(row=row_idx, column=c).fill = total_fill
        ws.cell(row=row_idx, column=c).border = border
    total_igv = sum(f["igv_redondeado"] for f in pendientes)
    cell_total = ws.cell(row=row_idx, column=6, value=total_igv)
    cell_total.font = total_font
    cell_total.number_format = "#,##0"
    cell_total.alignment = right

    ws.freeze_panes = "A2"
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

def generar_pdf_nps(filas: list, mes: str, empresa_nombre: str) -> bytes:
    """Genera un PDF simple con los datos necesarios para crear los NPS en SUNAT."""
    pendientes = filas_nps_pendientes(filas)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(letter),
        leftMargin=1.5 * cm, rightMargin=1.5 * cm, topMargin=1.5 * cm, bottomMargin=1.5 * cm
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Heading1"], fontSize=16, textColor=colors.HexColor("#1e3a5f"))
    subtitle_style = ParagraphStyle("subtitle", parent=styles["Normal"], fontSize=10, textColor=colors.HexColor("#64748b"))

    story = [
        Paragraph(f"IGV No Domiciliado — Pendiente de pago en SUNAT", title_style),
        Paragraph(f"{empresa_nombre} · Periodo {mes}", subtitle_style),
        Spacer(1, 16),
    ]

    table_data = [["Fecha", "Proveedor", "N° Invoice", "Monto USD", "IGV S/. a pagar"]]
    for f in pendientes:
        fecha_fmt = f["fecha"]
        try:
            fecha_fmt = datetime.strptime(f["fecha"], "%Y-%m-%d").strftime("%d/%m/%Y")
        except Exception:
            pass
        table_data.append([
            fecha_fmt, f["proveedor"], f["invoice"],
            f"$ {f['usd']:,.2f}", f"S/ {f['igv_redondeado']:,}"
        ])

    total_igv = sum(f["igv_redondeado"] for f in pendientes)
    table_data.append(["", "", "", "Total IGV a pagar:", f"S/ {total_igv:,}"])

    tbl = Table(table_data, colWidths=[3.2 * cm, 7.5 * cm, 5 * cm, 3.5 * cm, 4 * cm])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (3, 0), (4, -1), "RIGHT"),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("GRID", (0, 0), (-1, -2), 0.5, colors.HexColor("#cbd5e1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#f8fafc")]),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#dbeafe")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 14))
    story.append(Paragraph(
        f"Nota: se excluyen invoices con IGV ≤ S/ 1, ya que no requieren generación de NPS individual. "
        f"Total de invoices con NPS pendiente: {len(pendientes)}.",
        subtitle_style
    ))

    doc.build(story)
    return buffer.getvalue()

# ── UI Principal ──────────────────────────────────────────────────────────────
st.markdown("""
<div class="main-header">
    <h1>🧾 Facturas del Exterior</h1>
    <p>Procesamiento automático de invoices · IGV no domiciliado · Retención de renta</p>
</div>
""", unsafe_allow_html=True)

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### ⚙️ Configuración")
    
    empresa = st.text_input(
        "Nombre de empresa",
        value=st.session_state.empresa_activa,
        placeholder="Ej: ESSENTTA, SERPRESS...",
        help="Puedes cambiar esto para procesar otra empresa."
    )
    st.session_state.empresa_activa = empresa.upper().strip() if empresa else "EMPRESA"

    mes = st.text_input(
        "Mes / período",
        value=st.session_state.mes_activo,
        placeholder="Ej: ABRIL_2026, MAYO_2026...",
        help="Se usará como nombre de pestaña en el Excel."
    )
    st.session_state.mes_activo = mes.upper().strip() if mes else "MES_2026"

    st.divider()

    if st.session_state.filas:
        empresas_cargadas = list(dict.fromkeys(f["empresa"] for f in st.session_state.filas))
        st.markdown("**Datos en memoria:**")
        for emp in empresas_cargadas:
            n = sum(1 for f in st.session_state.filas if f["empresa"] == emp)
            st.markdown(f'<span class="empresa-tag">{emp}</span> · {n} invoices', unsafe_allow_html=True)
        
        st.divider()
        if st.button("🗑️ Limpiar todo", use_container_width=True):
            st.session_state.filas = []
            st.session_state.tipo_cambio = {}
            st.session_state.nps_cruzados = {}
            st.rerun()
    
    st.divider()
    st.markdown("**Cómo usar:**")
    st.markdown("""
1. Configura empresa y mes
2. Sube el PDF de tipo de cambio SUNAT
3. Sube los PDFs de invoices
4. Revisa y edita la tabla
5. Sube el reporte "Detalle de Declaraciones y Pagos" de SUNAT para llenar N° orden
6. Descarga el Excel
    """)

# ── Tabs principales ──────────────────────────────────────────────────────────
tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
    "📄 1. Tipo de cambio",
    "📥 2. Invoices",
    "✏️ 3. Revisar tabla",
    "🔢 4. N° Órdenes SUNAT",
    "📊 5. Exportar Excel",
    "💰 6. Renta por proveedor"
])

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 1 — Tipo de cambio
# ═══════════════════════════════════════════════════════════════════════════════
with tab1:
    st.markdown('<div class="step-card"><h3>📄 Paso 1 — Tipo de cambio SUNAT</h3>Sube el PDF mensual de tipo de cambio que descargas de SUNAT. La IA extrae todos los valores automáticamente.</div>', unsafe_allow_html=True)

    col1, col2 = st.columns([1, 1])
    with col1:
        pdf_tc = st.file_uploader("PDF Tipo de Cambio SUNAT", type=["pdf"], key="tc_uploader")
        
        if pdf_tc:
            with st.spinner("Extrayendo tipo de cambio..."):
                try:
                    tc_dict = extraer_tipo_cambio(pdf_tc.read())
                    st.session_state.tipo_cambio = tc_dict
                    st.markdown('<div class="success-box">✅ Tipo de cambio cargado correctamente.</div>', unsafe_allow_html=True)
                except Exception as e:
                    st.error(f"Error al procesar: {e}")

    with col2:
        if st.session_state.tipo_cambio:
            st.markdown("**Tipo de cambio venta cargado:**")
            tc_df = pd.DataFrame([
                {"Día": k, "TC Venta": v}
                for k, v in sorted(st.session_state.tipo_cambio.items())
            ])
            st.dataframe(tc_df, use_container_width=True, hide_index=True, height=300)
        else:
            st.markdown('<div class="info-box">ℹ️ Aún no hay tipo de cambio cargado. Sube el PDF de SUNAT.</div>', unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 2 — Invoices
# ═══════════════════════════════════════════════════════════════════════════════
with tab2:
    st.markdown('<div class="step-card"><h3>📥 Paso 2 — Subir invoices</h3>Sube todos los PDFs de facturas del mes. La IA extrae automáticamente los datos de cada uno.</div>', unsafe_allow_html=True)

    if not st.session_state.tipo_cambio:
        st.markdown('<div class="warning-box">⚠️ Primero carga el tipo de cambio SUNAT en el Paso 1.</div>', unsafe_allow_html=True)
    else:
        pdfs_invoices = st.file_uploader(
            "PDFs de invoices (puedes subir varios a la vez)",
            type=["pdf"],
            accept_multiple_files=True,
            key="invoices_uploader"
        )

        if pdfs_invoices:
            empresa_actual = st.session_state.empresa_activa
            st.info(f"Se procesarán {len(pdfs_invoices)} invoice(s) para la empresa **{empresa_actual}**")
            
            if st.button("🤖 Procesar invoices con IA", type="primary", use_container_width=True):
                progress = st.progress(0)
                status = st.empty()
                nuevos = 0
                
                for i, pdf_file in enumerate(pdfs_invoices):
                    status.markdown(f"Procesando **{pdf_file.name}**...")
                    result = extraer_invoice(
                        pdf_file.read(),
                        st.session_state.tipo_cambio,
                        empresa_actual
                    )
                    if result:
                        st.session_state.filas.append(result)
                        nuevos += 1
                    progress.progress((i + 1) / len(pdfs_invoices))
                
                status.empty()
                progress.empty()
                st.markdown(f'<div class="success-box">✅ Se procesaron <strong>{nuevos}</strong> invoice(s) correctamente. Ve al Paso 3 para revisar.</div>', unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 3 — Revisar tabla
# ═══════════════════════════════════════════════════════════════════════════════
with tab3:
    st.markdown('<div class="step-card"><h3>✏️ Paso 3 — Revisar y editar</h3>Verifica los datos extraídos. Puedes editar cualquier celda directamente.</div>', unsafe_allow_html=True)

    if not st.session_state.filas:
        st.markdown('<div class="info-box">ℹ️ Aún no hay datos. Procesa los invoices en el Paso 2.</div>', unsafe_allow_html=True)
    else:
        filas = st.session_state.filas
        
        # Métricas resumen
        empresas_unicas = list(dict.fromkeys(f["empresa"] for f in filas))
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Total invoices", len(filas))
        with col2:
            st.metric("Total USD", f"${sum(f['usd'] for f in filas):,.2f}")
        with col3:
            st.metric("Total S/. sin IGV", f"S/{sum(f['sin_igv'] for f in filas):,.2f}")
        with col4:
            st.metric("IGV total S/.", f"S/{sum(f['igv_soles'] for f in filas):,.2f}")

        st.divider()

        # Filtro por empresa
        if len(empresas_unicas) > 1:
            empresa_filtro = st.selectbox("Filtrar por empresa:", ["Todas"] + empresas_unicas)
        else:
            empresa_filtro = "Todas"

        filas_mostrar = filas if empresa_filtro == "Todas" else [f for f in filas if f["empresa"] == empresa_filtro]

        # Tabla editable
        df = pd.DataFrame(filas_mostrar)
        display_cols = ["empresa", "fecha", "proveedor", "pais", "vat", "invoice", "nac", "usd", "tc", "sin_igv", "igv_soles", "igv_usd", "total_soles", "renta", "renta_usd"]
        col_labels = {
            "empresa": "Empresa", "fecha": "Fecha", "proveedor": "Proveedor",
            "pais": "País", "vat": "RUC/VAT", "invoice": "N° Invoice",
            "nac": "Nacionalización", "usd": "USD", "tc": "T.C. Venta",
            "sin_igv": "S/. sin IGV", "igv_soles": "IGV S/.", "igv_usd": "IGV USD",
            "total_soles": "Total S/.", "renta": "Renta S/.", "renta_usd": "Renta USD"
        }
        df_display = df[display_cols].rename(columns=col_labels)
        
        edited_df = st.data_editor(
            df_display,
            use_container_width=True,
            num_rows="dynamic",
            hide_index=True,
            column_config={
                "USD": st.column_config.NumberColumn(format="%.2f"),
                "T.C. Venta": st.column_config.NumberColumn(format="%.3f"),
                "S/. sin IGV": st.column_config.NumberColumn(format="%.2f"),
                "IGV S/.": st.column_config.NumberColumn(format="%.2f"),
                "IGV USD": st.column_config.NumberColumn(format="%.2f"),
                "Total S/.": st.column_config.NumberColumn(format="%.2f"),
                "Renta S/.": st.column_config.NumberColumn(format="%.2f"),
                "Renta USD": st.column_config.NumberColumn(format="%.2f"),
            },
            key="tabla_editor"
        )

        if st.button("💾 Guardar cambios", type="primary"):
            inv_labels = {v: k for k, v in col_labels.items()}
            edited_back = edited_df.rename(columns=inv_labels)
            
            if empresa_filtro == "Todas":
                nuevas_filas = edited_back.to_dict("records")
            else:
                otras = [f for f in filas if f["empresa"] != empresa_filtro]
                nuevas_filas = otras + edited_back.to_dict("records")
            
            # Recalcular campos derivados si cambiaron USD o TC
            for fila in nuevas_filas:
                try:
                    usd = float(fila.get("usd", 0))
                    tc = float(fila.get("tc", 0))
                    calcs = calcular_fila(usd, tc)
                    fila.update(calcs)
                except:
                    pass
            
            st.session_state.filas = nuevas_filas
            st.success("✅ Cambios guardados.")
            st.rerun()

        st.divider()

        # ── Reporte de IGV / NPS pendientes (siempre visible, exportar es opcional) ──
        st.markdown('<div class="step-card"><h3>🧾 IGV para crear NPS en SUNAT</h3>Estos son los montos de IGV por invoice para crear las constancias NPS y enviarlas al cliente. Se excluyen automáticamente los invoices con IGV ≤ S/ 1.</div>', unsafe_allow_html=True)

        pendientes_preview = filas_nps_pendientes(filas_mostrar)
        excluidos = len(filas_mostrar) - len(pendientes_preview)

        c1, c2, c3 = st.columns(3)
        c1.metric("Invoices con NPS pendiente", len(pendientes_preview))
        c2.metric("Total IGV a pagar", f"S/ {sum(f['igv_redondeado'] for f in pendientes_preview):,}")
        c3.metric("Excluidos (IGV ≤ S/1)", excluidos)

        if pendientes_preview:
            tabla_nps = pd.DataFrame([
                {
                    "Fecha": f["fecha"],
                    "Proveedor": f["proveedor"],
                    "N° Invoice": f["invoice"],
                    "Monto USD": f["usd"],
                    "IGV S/. a pagar": f["igv_redondeado"],
                }
                for f in pendientes_preview
            ])
            st.dataframe(
                tabla_nps,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Monto USD": st.column_config.NumberColumn(format="%.2f"),
                    "IGV S/. a pagar": st.column_config.NumberColumn(format="%d"),
                },
            )
            st.caption("Puedes copiar estos montos directamente al crear cada NPS en SUNAT — no es necesario exportar nada para continuar.")
        else:
            st.markdown('<div class="info-box">ℹ️ No hay invoices con IGV mayor a S/1 en esta vista.</div>', unsafe_allow_html=True)

        with st.expander("⬇️ Exportar este reporte (opcional)"):
            col_exp1, col_exp2 = st.columns(2)
            empresa_nombre_export = empresa_filtro if empresa_filtro != "Todas" else "Todas las empresas"

            with col_exp1:
                excel_nps_bytes = generar_excel_nps(filas_mostrar, st.session_state.mes_activo or "MES")
                st.download_button(
                    "📊 Descargar Excel (NPS pendientes)",
                    data=excel_nps_bytes,
                    file_name=f"IGV_NPS_pendientes_{empresa_nombre_export}_{st.session_state.mes_activo or 'MES'}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            with col_exp2:
                pdf_nps_bytes = generar_pdf_nps(filas_mostrar, st.session_state.mes_activo or "MES", empresa_nombre_export)
                st.download_button(
                    "📄 Descargar PDF (NPS pendientes)",
                    data=pdf_nps_bytes,
                    file_name=f"IGV_NPS_pendientes_{empresa_nombre_export}_{st.session_state.mes_activo or 'MES'}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                )

        st.divider()

        # Agregar fila manual
        with st.expander("➕ Agregar invoice manualmente"):
            c1, c2, c3 = st.columns(3)
            with c1:
                m_empresa = st.text_input("Empresa", value=st.session_state.empresa_activa, key="m_emp")
                m_fecha = st.text_input("Fecha (YYYY-MM-DD)", key="m_fecha")
                m_proveedor = st.text_input("Proveedor", key="m_prov")
                m_pais = st.text_input("País", key="m_pais")
                m_domicilio = st.text_input("Domicilio", key="m_dom")
            with c2:
                m_vat = st.text_input("RUC/VAT", key="m_vat")
                m_invoice = st.text_input("N° Invoice", key="m_inv")
                m_usd = st.number_input("Monto USD", min_value=0.0, step=0.01, key="m_usd")
                m_tc = st.number_input("Tipo de cambio", min_value=0.0, step=0.001, format="%.3f", key="m_tc")
            with c3:
                if m_usd and m_tc:
                    calcs = calcular_fila(m_usd, m_tc)
                    st.metric("S/. sin IGV", f"{calcs['sin_igv']:,.2f}")
                    st.metric("IGV S/.", f"{calcs['igv_soles']:,.2f}")
                    st.metric("Total S/.", f"{calcs['total_soles']:,.2f}")
                    st.metric("Renta S/.", f"{calcs['renta']:,.2f}")
            
            if st.button("Agregar fila", key="btn_agregar"):
                if m_fecha and m_proveedor and m_usd:
                    calcs = calcular_fila(m_usd, m_tc)
                    st.session_state.filas.append({
                        "empresa": m_empresa.upper(),
                        "fecha": m_fecha,
                        "proveedor": m_proveedor,
                        "pais": m_pais,
                        "domicilio": m_domicilio,
                        "vat": m_vat,
                        "invoice": m_invoice,
                        "nac": "16630-",
                        "usd": m_usd,
                        "tc": m_tc,
                        **calcs,
                        "estado": "Manual"
                    })
                    st.success("✅ Fila agregada.")
                    st.rerun()

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 4 — N° Órdenes SUNAT
# ═══════════════════════════════════════════════════════════════════════════════
with tab4:
    st.markdown('<div class="step-card"><h3>🔢 Paso 4 — Asignar N° de órdenes SUNAT</h3>Elige el tipo de documento que tienes disponible para cruzar los pagos con los invoices.</div>', unsafe_allow_html=True)

    if not st.session_state.filas:
        st.markdown('<div class="info-box">ℹ️ Primero carga y revisa los invoices en los pasos anteriores.</div>', unsafe_allow_html=True)
    else:
        metodo = st.radio(
            "¿Qué documento vas a subir?",
            ["📋 Formulario 1662 — Resumen de Transacciones (método nuevo)",
             "📊 Detalle de Declaraciones y Pagos — Excel/PDF (método anterior)"],
            key="metodo_cruce"
        )

        if "1662" in metodo:
            st.markdown('<div class="info-box">ℹ️ Sube el PDF "Resumen de Transacciones" que genera SUNAT al pagar con el Formulario 1662. El sistema cruza usando el número de comprobante de cada pago (que coincide con el N° de invoice).</div>', unsafe_allow_html=True)
            archivo_1662 = st.file_uploader(
                "PDF Resumen de Transacciones (Formulario 1662)",
                type=["pdf"], key="uploader_1662"
            )
            if archivo_1662:
                if st.button("🔄 Cruzar con Formulario 1662", type="primary", use_container_width=True):
                    with st.spinner("Leyendo Formulario 1662..."):
                        try:
                            pagos_1662 = extraer_formulario_1662(archivo_1662.read())
                        except Exception as e:
                            st.error(f"Error: {e}")
                            pagos_1662 = []

                    if not pagos_1662:
                        st.markdown('<div class="warning-box">⚠️ No se encontraron pagos en el PDF.</div>', unsafe_allow_html=True)
                    else:
                        cruce = cruzar_formulario_1662_con_invoices(pagos_1662, st.session_state.filas)
                        for idx, n_orden in cruce.items():
                            st.session_state.filas[idx]["nac"] = f"16630-{n_orden}"
                        st.session_state.nps_cruzados = cruce

                        st.markdown(f'<div class="success-box">✅ Se cruzaron <strong>{len(cruce)}</strong> de {len(st.session_state.filas)} invoices con número de orden.</div>', unsafe_allow_html=True)

                        st.markdown("**Pagos detectados:**")
                        st.dataframe(pd.DataFrame(pagos_1662), use_container_width=True, hide_index=True)

                        if cruce:
                            st.markdown("**Resultado del cruce:**")
                            resumen = []
                            for idx, n_orden in cruce.items():
                                fila = st.session_state.filas[idx]
                                resumen.append({
                                    "Proveedor": fila["proveedor"],
                                    "Invoice": fila["invoice"],
                                    "N° Orden asignado": f"16630-{n_orden}"
                                })
                            st.dataframe(pd.DataFrame(resumen), use_container_width=True, hide_index=True)

        else:
            st.markdown('<div class="info-box">ℹ️ Descarga el reporte desde SUNAT Operaciones en Línea → Consultas → Detalle de Declaraciones y Pagos. Acepta PDF y Excel (.xls/.xlsx).</div>', unsafe_allow_html=True)
            archivo_detalle = st.file_uploader(
                "Reporte 'Detalle de Declaraciones y Pagos' (PDF o Excel)",
                type=["pdf", "xls", "xlsx"], key="detalle_uploader"
            )
            if archivo_detalle:
                if st.button("🔄 Cruzar pagos con invoices", type="primary", use_container_width=True):
                    with st.spinner("Leyendo reporte SUNAT..."):
                        try:
                            pagos = extraer_detalle_declaraciones(archivo_detalle.read(), archivo_detalle.name)
                        except Exception as e:
                            st.error(f"Error al leer el archivo: {e}")
                            pagos = []

                    if not pagos:
                        st.markdown('<div class="warning-box">⚠️ No se encontraron pagos tipo "BOLETA DE PAGO - NPS" en el archivo.</div>', unsafe_allow_html=True)
                    else:
                        cruce = cruzar_pagos_con_invoices(pagos, st.session_state.filas)
                        for idx, n_orden in cruce.items():
                            st.session_state.filas[idx]["nac"] = f"16630-{n_orden}"
                        st.session_state.nps_cruzados = cruce

                        st.markdown(f'<div class="success-box">✅ Se cruzaron <strong>{len(cruce)}</strong> de {len(st.session_state.filas)} invoices con número de orden.</div>', unsafe_allow_html=True)
                        st.markdown("**Pagos detectados:**")
                        st.dataframe(pd.DataFrame(pagos), use_container_width=True, hide_index=True)

                        if cruce:
                            st.markdown("**Resultado del cruce:**")
                            resumen = []
                            for idx, n_orden in cruce.items():
                                fila = st.session_state.filas[idx]
                                resumen.append({
                                    "Proveedor": fila["proveedor"],
                                    "Invoice": fila["invoice"],
                                    "IGV S/. (redondeado)": int(round(fila["igv_soles"])),
                                    "N° Orden asignado": f"16630-{n_orden}"
                                })
                            st.dataframe(pd.DataFrame(resumen), use_container_width=True, hide_index=True)

        sin_cruzar = [i for i in range(len(st.session_state.filas)) if i not in st.session_state.nps_cruzados]
        if sin_cruzar:
            st.markdown("**⚠️ Invoices sin N° de orden:**")
            for idx in sin_cruzar:
                f = st.session_state.filas[idx]
                st.markdown(f"- {f['proveedor']} · {f['invoice']} · IGV S/. {int(round(f['igv_soles']))}")

    st.divider()

    # ── Asignación manual de N° de orden ─────────────────────────────────────

    st.markdown('<div class="step-card"><h3>✏️ O ingresa los N° de orden manualmente</h3>Si ya tienes los números de orden a mano (por ejemplo porque ya entraste a SUNAT a revisarlos), puedes escribirlos directamente aquí sin necesidad de subir ningún archivo.</div>', unsafe_allow_html=True)

    if not st.session_state.filas:
        st.markdown('<div class="info-box">ℹ️ Aún no hay invoices cargados. Puedes agregarlos en el Paso 2, o agregar uno manualmente en el Paso 3.</div>', unsafe_allow_html=True)
    else:
        tabla_manual = pd.DataFrame([
            {
                "Empresa": f["empresa"],
                "Proveedor": f["proveedor"],
                "N° Invoice": f["invoice"],
                "IGV S/.": int(round(f["igv_soles"])),
                "N° de orden (sin 16630-)": f.get("nac", "16630-").replace("16630-", ""),
            }
            for f in st.session_state.filas
        ])

        tabla_editada = st.data_editor(
            tabla_manual,
            use_container_width=True,
            hide_index=True,
            disabled=["Empresa", "Proveedor", "N° Invoice", "IGV S/."],
            key="editor_nro_orden",
        )

        if st.button("💾 Guardar números de orden", type="primary", use_container_width=True):
            for idx, row in tabla_editada.iterrows():
                n_orden = str(row["N° de orden (sin 16630-)"]).strip()
                st.session_state.filas[idx]["nac"] = f"16630-{n_orden}" if n_orden else "16630-"
            st.success("✅ Números de orden guardados.")
            st.rerun()

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 5 — Exportar
# ═══════════════════════════════════════════════════════════════════════════════
with tab5:
    st.markdown('<div class="step-card"><h3>📊 Paso 5 — Exportar Excel</h3>Genera el Excel final con una pestaña por empresa, en el formato exacto de tu plantilla.</div>', unsafe_allow_html=True)

    if not st.session_state.filas:
        st.markdown('<div class="info-box">ℹ️ No hay datos para exportar. Completa los pasos anteriores.</div>', unsafe_allow_html=True)
    else:
        filas = st.session_state.filas
        empresas = list(dict.fromkeys(f["empresa"] for f in filas))
        mes = st.session_state.mes_activo or "MES_2026"

        # Resumen por empresa
        st.markdown("**Resumen por empresa:**")
        for emp in empresas:
            rows_emp = [f for f in filas if f["empresa"] == emp]
            tot_usd = sum(f["usd"] for f in rows_emp)
            tot_soles = sum(f["sin_igv"] for f in rows_emp)
            tot_igv = sum(f["igv_soles"] for f in rows_emp)
            tot_renta = sum(f["renta"] for f in rows_emp)
            sin_orden = sum(1 for f in rows_emp if f.get("nac", "16630-") == "16630-")
            
            with st.expander(f"🏢 {emp} · {len(rows_emp)} invoices", expanded=True):
                c1, c2, c3, c4 = st.columns(4)
                c1.metric("Total USD", f"${tot_usd:,.2f}")
                c2.metric("Total S/. sin IGV", f"S/{tot_soles:,.2f}")
                c3.metric("IGV Total S/.", f"S/{tot_igv:,.2f}")
                c4.metric("Retención Renta S/.", f"S/{tot_renta:,.2f}")
                if sin_orden > 0:
                    st.markdown(f'<div class="warning-box">⚠️ {sin_orden} invoice(s) aún sin número de orden (Nacionalización = "16630-"). Puedes descargar igual y completar después.</div>', unsafe_allow_html=True)

        st.divider()

        nombre_archivo = f"{'_'.join(empresas)}_{mes}.xlsx"
        
        if st.button("📥 Generar y descargar Excel", type="primary", use_container_width=True):
            with st.spinner("Generando Excel..."):
                excel_bytes = generar_excel(filas, mes)
            
            st.download_button(
                label="⬇️ Descargar Excel ahora",
                data=excel_bytes,
                file_name=nombre_archivo,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            st.markdown(f'<div class="success-box">✅ Excel generado: <strong>{nombre_archivo}</strong><br>Pestañas incluidas: {", ".join(f"{e}_{mes}" for e in empresas)}</div>', unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 6 — Renta por proveedor
# ═══════════════════════════════════════════════════════════════════════════════
with tab6:
    st.markdown('<div class="step-card"><h3>💰 Retención de Renta por proveedor</h3>Resumen del total S/. sin IGV agrupado por proveedor, con el cálculo del 30% de retención de renta para declarar en SUNAT. Ingresas cada proveedor con su monto base y SUNAT calcula el 30% automáticamente.</div>', unsafe_allow_html=True)

    if not st.session_state.filas:
        st.markdown('<div class="info-box">ℹ️ Carga los invoices del mes en el Paso 2 para ver este resumen.</div>', unsafe_allow_html=True)
    else:
        filas_renta = st.session_state.filas
        empresa_renta = st.selectbox(
            "Filtrar por empresa:",
            ["Todas"] + list(dict.fromkeys(f["empresa"] for f in filas_renta)),
            key="sel_empresa_renta"
        )
        if empresa_renta != "Todas":
            filas_renta = [f for f in filas_renta if f["empresa"] == empresa_renta]

        renta_data = calcular_renta_por_proveedor(filas_renta)

        # Métricas
        total_base = sum(r["Total S/. sin IGV"] for r in renta_data)
        total_renta = sum(r["Renta 30% S/."] for r in renta_data)
        col1, col2, col3 = st.columns(3)
        col1.metric("Proveedores", len(renta_data))
        col2.metric("Total base S/. sin IGV", f"S/ {total_base:,.2f}")
        col3.metric("Total Renta 30% S/.", f"S/ {total_renta:,.2f}")

        st.markdown("**Detalle por proveedor — para ingresar en SUNAT:**")
        st.caption("Ingresa cada proveedor con su 'Total S/. sin IGV' en el formulario de SUNAT. SUNAT calculará automáticamente el 30% de retención.")

        df_renta = pd.DataFrame(renta_data)
        st.dataframe(
            df_renta,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Total S/. sin IGV": st.column_config.NumberColumn(format="%.2f"),
                "Renta 30% S/.": st.column_config.NumberColumn(format="%.2f"),
                "N° invoices": st.column_config.NumberColumn(format="%d"),
            }
        )

        st.markdown('<div class="info-box">ℹ️ La pestaña <strong>RENTA_{mes}</strong> en el Excel exportado (Paso 5) contiene este mismo resumen.</div>', unsafe_allow_html=True)
